#!/usr/bin/env python3
"""
Skript pro konverzi English.xlsx na cards.json
Spustit kdykoli přidáte nové věty do XLSX.
"""

import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instaluji openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl


# Cesta k XLSX souboru na Google Drive (pouze pro čtení)
XLSX_PATH = Path(r"G:\Můj disk\Gdisk\English.xlsx")
OUTPUT_PATH = Path(__file__).parent / "cards.json"


def load_xlsx(xlsx_path: Path) -> dict:
    """Načte XLSX a vrátí strukturu pro JSON."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    cards = []

    # Pouze záložka "Gramatika věty"
    sheet_name = "Gramatika věty"
    if sheet_name not in wb.sheetnames:
        print(f"CHYBA: Záložka '{sheet_name}' nenalezena!")
        print(f"Dostupné záložky: {wb.sheetnames}")
        return {"cards": []}

    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=1, values_only=True):
        # Přeskočíme prázdné řádky
        if not row or len(row) < 2:
            continue

        # Sloupec B = anglicky, sloupec C = česky
        en = row[1] if len(row) > 1 else None
        cz = row[2] if len(row) > 2 else None

        # Přeskočíme pokud chybí EN nebo CZ
        if not en or not cz:
            continue

        # Přeskočíme "nadpisy" (řádky bez překladu)
        en_str = str(en).strip()
        cz_str = str(cz).strip()

        if not en_str or not cz_str:
            continue

        cards.append({
            "en": en_str,
            "cz": cz_str,
            "category": sheet_name
        })

    wb.close()
    return {"cards": cards}


def main():
    if not XLSX_PATH.exists():
        print(f"CHYBA: Soubor {XLSX_PATH} neexistuje!")
        print("Upravte XLSX_PATH v tomto skriptu.")
        sys.exit(1)

    print(f"Načítám: {XLSX_PATH}")
    data = load_xlsx(XLSX_PATH)

    print(f"Nalezeno {len(data['cards'])} kartiček")

    # Uložení do JSON
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Uloženo do: {OUTPUT_PATH}")

    # Spustit server a otevřít prohlížeč
    import http.server
    import webbrowser
    import threading

    PORT = 8080
    PROJECT_DIR = OUTPUT_PATH.parent
    PROGRESS_FILE = PROJECT_DIR / "progress.json"

    class MyHandler(http.server.SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=str(PROJECT_DIR), **kwargs)

        def do_POST(self):
            if self.path == "/save-progress":
                content_length = int(self.headers["Content-Length"])
                post_data = self.rfile.read(content_length)

                with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
                    f.write(post_data.decode("utf-8"))

                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Access-Control-Allow-Origin", "*")
                self.end_headers()
                self.wfile.write(b'{"status": "ok"}')
            else:
                self.send_response(404)
                self.end_headers()

        def do_OPTIONS(self):
            self.send_response(200)
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Methods", "POST, GET, OPTIONS")
            self.send_header("Access-Control-Allow-Headers", "Content-Type")
            self.end_headers()

    with http.server.HTTPServer(("", PORT), MyHandler) as httpd:
        url = f"http://localhost:{PORT}"
        print(f"\nServer běží na: {url}")
        print("Pokrok se ukládá do: {PROGRESS_FILE}")
        print("Pro ukončení stiskni Ctrl+C")

        # Otevřít prohlížeč
        threading.Timer(0.5, lambda: webbrowser.open(url)).start()

        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\nServer ukončen.")


if __name__ == "__main__":
    main()
